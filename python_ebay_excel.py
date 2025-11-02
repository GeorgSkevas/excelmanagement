import openpyxl
from datetime import datetime, date
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
import os


#COLORS
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
gray_fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")




name_of_sheet = input("Please enter the name of the excel sheet (with the ending .xlsm): ")

# Load workbook
wb = openpyxl.load_workbook(name_of_sheet, keep_vba=True)


#set sheet's variables
endresult_sheet = wb['end result']
jtl_sheet = wb['jtl']
ebay_sheet = wb['ebay']
retouren_sheet = wb['retouren']






# BOLD BOX
thick_border = Border(
    left = Side(border_style="thin", color="FF000000"),
    right = Side(border_style='thin', color="FF000000"),
    top = Side(border_style="thin", color="FF000000"),
    bottom = Side(border_style="thin", color="FF000000")
)


# Booking codes in JTL sheet
jtl_dict = {}

for j in range(1,3):
    i0 = 2
    while True:
        date_jtl_counter = jtl_sheet.cell(row=i0, column=6).value
        if date_jtl_counter is None:
            break

        key = jtl_sheet.cell(row=i0, column=j).value
        jtl_dict[key] = {
        'order_nr': jtl_sheet.cell(row=i0, column=7).value,
        'buchungskonto': jtl_sheet.cell(row=i0, column=18).value,
        'vorname': jtl_sheet.cell(row=i0, column=3).value,
        'nachname': jtl_sheet.cell(row=i0, column=4).value
        }

        i0 += 1

print(f'Number of JTL sheet`s lines: {i0}')



# Booking codes in Retouren Sheet
i1 = 2
retouren_dict = {}

while True:
    retouren_counter = retouren_sheet.cell(row=i1, column=2).value
    if retouren_counter is None:
        break

    key = retouren_sheet.cell(row=i1, column=1).value
    retouren_dict[key] = {
                    'order_nr': retouren_sheet.cell(row=i1, column=5).value,
                    'buchungskonto': retouren_sheet.cell(row=i1, column=4).value,
                    'name': retouren_sheet.cell(row=i1, column=7).value, 
    }
    i1+=1

print(f'Number of RETOUREN sheet´s lines: {i1}')
 
                       
def ebay_sort(endresult_sheet, ebay_sheet):
    #Titles
    endresult_sheet.cell(row=3, column=6).value = "Datum"
    endresult_sheet.cell(row=3, column=7).value = "Text"
    endresult_sheet.cell(row=3, column=4).value = "Rechn-Nr"
    endresult_sheet.cell(row=3, column=5).value = "Au-Nr"
    endresult_sheet.cell(row=3, column=8).value = "Gefundedn in"
    endresult_sheet.cell(row=3, column=3).value = "GegenKto"
    endresult_sheet.cell(row=3, column=1).value = "Einnahmen"
    endresult_sheet.cell(row=3, column=2).value = "Ausgaben"
    for bold in range (1,9):
        endresult_sheet.cell(row=3, column=bold).font = Font(bold=True)
        
  

    #Datum
    dates = []

    # counter
    i = 2
    dates_counter = ebay_sheet.cell(row=i, column=1).value

    while dates_counter != None:
        cell_value = ebay_sheet.cell(row=i, column=1).value
        if isinstance(cell_value, datetime):
            cell_value = cell_value.date()
        elif isinstance(cell_value, str):
            try:
                cell_value = datetime.strptime(cell_value.strip(), "%d.%m.%Y").date()              
            except:
                continue 

        cell = endresult_sheet.cell(row=i+2, column=6)
        cell.value = cell_value
        cell.number_format = 'DD.MM.YYYY'

        if isinstance(cell_value, date):
            dates.append(cell_value)


        else:
            endresult_sheet.cell(row=4, column=11).value = "No valid dates"

        # set variables from ebay sheet, transactions names and booking codes 
        transaction_type= ebay_sheet.cell(row=i, column=2).value
        name_cell = ebay_sheet.cell(row=i, column=6).value
        booking_code = ebay_sheet.cell(row=i, column=3).value


        # add the filter in the titles
        endresult_sheet.auto_filter.ref = "A3:H3"


        # make the border thick
        for col in range(1,9):
            cell = endresult_sheet.cell(row=3, column=col)
            cell.border = thick_border
            cell = endresult_sheet.cell(row=i+2, column=col)
            cell.border = thick_border


        #write the name in end result
        endresult_sheet.cell(row=i+2, column=7).value = name_cell

        #search order nr buchungskonto in JTL and RETOUREN  

        if booking_code != None:
            if booking_code in jtl_dict:
                jtl_data = jtl_dict[booking_code]
                full_name = f"{jtl_data['vorname']} {jtl_data['nachname']}".strip()
                endresult_sheet.cell(row=i+2, column=7).value = full_name            
                endresult_sheet.cell(row=i+2, column=4).value = jtl_data['order_nr']
                endresult_sheet.cell(row=i+2,column=3).value = jtl_data['buchungskonto']
                endresult_sheet.cell(row=i+2, column=8).value = "gefunden in JTL"
            elif booking_code in retouren_dict:
                retouren_data = retouren_dict[booking_code]
                endresult_sheet.cell(row=i+2, column=7).value = retouren_data['name']
                endresult_sheet.cell(row=i+2, column=4).value = retouren_data['order_nr']
                endresult_sheet.cell(row=i+2,column=3).value = retouren_data['buchungskonto']
                endresult_sheet.cell(row=i+2, column=8).value = "gefunden in RETOUREN"
            else:
                not_found = "Nicht gefunden"
                endresult_sheet.cell(row=i+2, column=8).value = not_found
                endresult_sheet.cell(row=i+2, column=7).fill = yellow_fill
                endresult_sheet.cell(row=i+2, column=7).value = transaction_type
                endresult_sheet.cell(row=i+2, column=5).value = booking_code


        if booking_code == None:
            endresult_sheet.cell(row=i+2, column=7).value = transaction_type
            endresult_sheet.cell(row=i+2, column=7).fill = red_fill
            endresult_sheet.cell(row=i+2, column=8).value = "Kein AU-Nr"



        # if the transaction is not Bestellung then change functionality of name's cell
        if transaction_type == "Andere Gebühr":
            endresult_sheet.cell(row=i+2, column=3).value = '4600'
        elif transaction_type == "Einbehalten":
            endresult_sheet.cell(row=i+2, column=3).value = '1590'
        elif transaction_type == "Fall":
            endresult_sheet.cell(row=i+2, column=3).value = "1590"
        if name_cell == '--' or booking_code == '--':
            transaction_type_ebay = ebay_sheet.cell(row=i, column=2).value
            endresult_sheet.cell(row=i+2, column=7).value = str(transaction_type_ebay) + ' : ' + str(name_cell)
            endresult_sheet.cell(row=i+2, column=7).fill = red_fill
            endresult_sheet.cell(row=i+2, column=8).value = 'No AU-Nr.'
        elif transaction_type != "Bestellung":
            transaction_type_ebay = ebay_sheet.cell(row=i, column=2).value
            endresult_sheet.cell(row=i+2, column=7).value = str(transaction_type_ebay) + ' : ' + str(name_cell)
            endresult_sheet.cell(row=i+2, column=7).fill = yellow_fill





        




        # Transaktionsbetrag     
        subtotal_item = ebay_sheet.cell(row=i, column=34).value

        if subtotal_item is not None:
            subtotal_str = str(subtotal_item).strip()

            if '.' in subtotal_str:
                parts = subtotal_str.split('.')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    cleaned_value = subtotal_str.replace('.', ',')
                else:
                    cleaned_value = subtotal_str
            else:
                cleaned_value = subtotal_str

            if cleaned_value.startswith('-'):
                endresult_sheet.cell(row=i+2, column=2).value = cleaned_value.lstrip('-')
            else:
                endresult_sheet.cell(row=i+2, column=1).value = cleaned_value





        
        i += 1
        dates_counter = ebay_sheet.cell(row=i, column=1).value
        # end of for loop 


    # print number of lines

    print("The number of ebay´s sheet lines are: ", i)



            # find min and max Datum

    if dates:
        min_date = min(dates).strftime("%d.%m.%Y")
        max_date = max(dates).strftime("%d.%m.%Y")
        for d in range(3, 5):
            for co in range(10, 12):
                endresult_sheet.cell(row=d, column=co).fill = gray_fill
                endresult_sheet.cell(row=d, column=co).border = thick_border
            endresult_sheet.cell(row=3, column=10).value = "Startdatum:"
            endresult_sheet.cell(row=4, column=10).value = "Enddatum:"    
            endresult_sheet.cell(row=3, column=11).value = min_date
            endresult_sheet.cell(row=4, column=11).value = max_date






        # write what the colors mean

    for instr in range(8, 11):
        for col_instr in range(10,12):
            endresult_sheet.cell(row=instr, column=col_instr).border = thick_border
    endresult_sheet.cell(row=8, column=10).fill = red_fill
    endresult_sheet.cell(row=9, column=11).value = "Andere Transaktion, mit oder ohne Ordernummer"
    endresult_sheet.cell(row=9, column=10).fill = yellow_fill
    endresult_sheet.cell(row=8, column=11).value = "Keine Kunden Transaktion oder Ordernummer"
    endresult_sheet.cell(row=10, column=11).value = "Ordernummer in JTL oder Retouren gefunden"



# end of function




def main():
    ebay_sort(endresult_sheet, ebay_sheet)



main()


def is_file_open01(filepath):
    try:
        # Try to open file for exclusive access
        with open(filepath, "a"):
            return False  # file is not locked
    except PermissionError:
        return True  # file is locked (likely open in Excel)

if is_file_open01(name_of_sheet):
    print(f"The file {name_of_sheet} is currently open. Please close it and rerun the script.")
    exit()


wb.save(name_of_sheet)

#width of cells
list = ['A','B','C','D','E','F','G','H','K']
for w in list:
    max_length = 0
    for row in range(2, endresult_sheet.max_row + 1):
        cell_value = endresult_sheet[f'{w}{row}'].value  # Reference cell by column letter and row number
        if cell_value:
            max_length = max(max_length, len(str(cell_value)))  # Find the longest string in the column
        else:
            continue
            
    
    # Set the width of the column
    endresult_sheet.column_dimensions[w].width = max_length + 3  # Add some 



# Speichern
def is_file_open02(filepath):
    try:
        # Try to open file for exclusive access
        with open(filepath, "a"):
            return False  # file is not locked
    except PermissionError:
        return True  # file is locked (likely open in Excel)

# Use it before running anything
if is_file_open02(name_of_sheet):
    print(f"The file {name_of_sheet} is currently open. Please close it and rerun the script.")
    exit()

wb.save(name_of_sheet)
print('\n\nSuccessfully Saved.\n\n')


input("Press enter to close the window: ")